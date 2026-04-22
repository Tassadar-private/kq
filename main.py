"""
Pobiera dane z Elektronicznych Ksiąg Wieczystych (EKW) do pliku ziemniak.xlsx.

Użycie:
    1. W pliku ziemniak.xlsx w kolumnie "Nr KW" wpisz pełne numery KW (np. WA1M/00123456/7).
    2. Uruchom:  python main.py
    3. Dla każdego numeru KW skrypt wypełni formularz na https://ekw.ms.gov.pl/eukw_ogol/menu.do
       i zatrzyma się — wtedy RĘCZNIE przepisz CAPTCHA i kliknij "Szukaj księgi",
       a następnie wróć do terminala i naciśnij ENTER.
    4. Skrypt pobierze dane z działu I-O i zapisze je w odpowiednich kolumnach.
"""

import os
import random
import re
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

EKW_URL = "https://ekw.ms.gov.pl/eukw_ogol/menu.do"

# --- prewencja przed blokadą IP przez EKW -----------------------------------
MIN_DELAY_SEC = 8          # minimalna losowa przerwa między zapytaniami
MAX_DELAY_SEC = 18         # maksymalna losowa przerwa między zapytaniami
BATCH_SIZE = 15            # co ile zapytań wymuszamy dłuższą przerwę
BATCH_PAUSE_SEC = 180      # długość dłuższej przerwy (sekundy)
HARD_LIMIT_PER_RUN = 80    # twardy limit zapytań na jedno uruchomienie skryptu

# Frazy, które EKW pokazuje przy przekroczeniu limitu / blokadzie IP.
BLOCK_PHRASES = (
    "przekroczono dopuszczaln",
    "zbyt wiele",
    "zablokowan",
    "limit zapyta",
    "chwilowo niedost",
    "prosimy spr",
)
# ---------------------------------------------------------------------------

SEARCH_DIRS = [
    Path.home() / "Desktop",
    Path.home() / "Pulpit",
    Path(__file__).parent,
    Path.home() / "OneDrive" / "Desktop",
    Path.home() / "OneDrive" / "Pulpit",
    Path.home() / "Documents",
    Path.home() / "Dokumenty",
    Path.home() / "Downloads",
    Path.home() / "Pobrane",
    Path.home(),
]

# Opcje ustawiane przed startem (menu w ask_run_options()).
OPTIONS = {
    "auto_focus_captcha": True,   # po wypełnieniu formularza ustaw kursor w polu CAPTCHA
    "beep_on_ready": True,        # dźwięk, gdy skrypt czeka na Twoją akcję
    "auto_detect_search": True,   # sam wykryj kliknięcie 'Szukaj księgi' (bez ENTER w terminalu)
}

COLUMNS = [
    "Nr KW", "Typ Księgi", "Stan Księgi", "Województwo", "Powiat", "Gmina",
    "Miejscowość", "Dzielnica", "Położenie", "Nr działek po średniku",
    "Obręb po średniku", "Ulica", "Sposób korzystania", "Obszar",
    "Ulica(dla budynku)", "przeznaczenie (dla budynku)",
    "Ulica(dla lokalu)", "Nr budynku( dla lokalu)",
    "Przeznaczenie (dla lokalu)", "Cały adres (dla lokalu)", "Czy udziały?",
]

KW_RE = re.compile(r"^\s*([A-Z0-9]{4})\s*/\s*(\d{8})\s*/\s*(\d)\s*$")


def find_existing_file() -> Path | None:
    """Szuka pliku ziemniak.xlsx w typowych lokalizacjach."""
    for d in SEARCH_DIRS:
        if not d.exists():
            continue
        for name in ("ziemniak.xlsx", "Ziemniak.xlsx", "ZIEMNIAK.xlsx"):
            p = d / name
            if p.is_file():
                return p
    return None


def prompt_yes(question: str, default: bool = True) -> bool:
    suffix = " [T/n]: " if default else " [t/N]: "
    ans = input(question + suffix).strip().lower()
    if not ans:
        return default
    return ans in ("t", "tak", "y", "yes")


def pick_file_dialog(initial_dir: Path | None = None) -> Path | None:
    """Otwiera natywne okno wyboru pliku (tkinter). Zwraca None, jeśli anulowano."""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception as exc:
        print(f"Nie mogę uruchomić okna dialogowego ({exc}). Podaj ścieżkę ręcznie.")
        return None
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path_str = filedialog.askopenfilename(
        title="Wybierz plik Excel (ziemniak.xlsx)",
        initialdir=str(initial_dir) if initial_dir else str(Path.home()),
        filetypes=[("Excel", "*.xlsx *.xlsm"), ("Wszystkie pliki", "*.*")],
    )
    root.destroy()
    if not path_str:
        return None
    p = Path(path_str)
    return p if p.is_file() else None


def desktop_dir() -> Path:
    """Zwraca istniejący katalog Pulpitu (PL/EN, także OneDrive)."""
    for d in (
        Path.home() / "Desktop",
        Path.home() / "Pulpit",
        Path.home() / "OneDrive" / "Desktop",
        Path.home() / "OneDrive" / "Pulpit",
    ):
        if d.is_dir():
            return d
    return Path.home()


def create_template(target: Path | None = None) -> Path:
    if target is None:
        target = desktop_dir() / "ziemniak.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "ziemniak"
    ws.append(COLUMNS)
    wb.save(target)
    print(f"Utworzono pusty szablon: {target}")
    print("Uzupełnij kolumnę 'Nr KW' numerami KW w formacie XXXX/NNNNNNNN/C.")
    return target


def locate_or_create_workbook() -> Path:
    """Najpierw szuka ziemniak.xlsx na Pulpicie. Jeśli jest — tylko potwierdza.
    Jeśli nie ma — pokazuje menu (okno dialogowe / ścieżka / szablon)."""
    found = find_existing_file()
    if found is not None:
        print(f"\nZnaleziono plik: {found}")
        print("(z tego pliku zostaną odczytane numery KW — w kolumnie 'Nr KW' —")
        print(" i do jego wierszy zostaną dopisane pobrane z EKW dane).")
        if prompt_yes("Uruchomić scrapera na tym pliku?"):
            return found
        print("OK, pokazuję menu wyboru.\n")

    print("\nSkąd wziąć plik Excel z numerami KW?")
    print("  [1] Okno wyboru pliku (polecane)")
    print("  [2] Znajdź automatycznie 'ziemniak.xlsx' w typowych lokalizacjach")
    print("  [3] Podam ścieżkę ręcznie")
    print("  [4] Utwórz nowy, pusty szablon (na Pulpicie)")
    choice = input("Wybór [1]: ").strip() or "1"

    if choice == "1":
        initial = find_existing_file()
        initial_dir = initial.parent if initial else None
        picked = pick_file_dialog(initial_dir)
        if picked is None:
            print("Nie wybrano pliku. Anulowano.")
            sys.exit(0)
        print(f"Wybrano: {picked}")
        if prompt_yes("Uruchomić scrapera na tym pliku?"):
            return picked
        sys.exit(0)

    if choice == "2":
        found = find_existing_file()
        if found is None:
            print("Nie znaleziono pliku 'ziemniak.xlsx' w typowych lokalizacjach.")
            sys.exit(1)
        print(f"Znaleziono: {found}")
        if prompt_yes("Uruchomić scrapera na tym pliku?"):
            return found
        sys.exit(0)

    if choice == "3":
        custom = input("Ścieżka do pliku: ").strip().strip('"')
        if not custom:
            print("Pusta ścieżka. Anulowano.")
            sys.exit(1)
        p = Path(custom).expanduser()
        if not p.is_file():
            print(f"Plik nie istnieje: {p}")
            sys.exit(1)
        if prompt_yes(f"Uruchomić scrapera na pliku {p}?"):
            return p
        sys.exit(0)

    if choice == "4":
        target = create_template()
        if not prompt_yes("Uruchomić scrapera teraz?", default=False):
            sys.exit(0)
        return target

    print("Nieznany wybór.")
    sys.exit(1)


def split_kw(kw: str) -> tuple[str, str, str] | None:
    m = KW_RE.match(kw)
    if not m:
        return None
    return m.group(1).upper(), m.group(2), m.group(3)


DEBUG_PORT = 9222
CHROME_PROFILE_DIR = Path(os.environ.get("LOCALAPPDATA", Path.home())) / "ekw_chrome_profile"

CHROME_CANDIDATES = [
    Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "Google/Chrome/Application/chrome.exe",
    Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")) / "Google/Chrome/Application/chrome.exe",
    Path(os.environ.get("LOCALAPPDATA", "")) / "Google/Chrome/Application/chrome.exe",
]


def find_chrome_exe() -> Path | None:
    for p in CHROME_CANDIDATES:
        if p and p.is_file():
            return p
    return None


def is_debug_chrome_running(port: int = DEBUG_PORT) -> bool:
    """Sprawdza, czy na lokalnym porcie działa Chrome z DevTools Protocol."""
    import socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.4)
        try:
            return s.connect_ex(("127.0.0.1", port)) == 0
        except OSError:
            return False


def launch_debug_chrome(port: int = DEBUG_PORT, url: str = EKW_URL) -> None:
    """Uruchamia Chrome w osobnym profilu z włączonym DevTools Protocol.
    Flagi ukrywają paski 'nieobsługiwana flaga' i 'sterowane przez automat'."""
    import subprocess

    if is_debug_chrome_running(port):
        print(f"Chrome z DevTools już nasłuchuje na porcie {port} — używam istniejącego.")
        return

    chrome = find_chrome_exe()
    if chrome is None:
        print("Nie znaleziono chrome.exe. Zainstaluj Google Chrome.")
        sys.exit(1)

    CHROME_PROFILE_DIR.mkdir(parents=True, exist_ok=True)
    args = [
        str(chrome),
        f"--remote-debugging-port={port}",
        f"--user-data-dir={CHROME_PROFILE_DIR}",
        "--test-type",
        "--no-first-run",
        "--no-default-browser-check",
        "--disable-infobars",
        "--disable-features=InfobarOnTranslate,Translate",
        "--disable-blink-features=AutomationControlled",
        url,
    ]
    print(f"Uruchamiam Chrome (port {port}, profil: {CHROME_PROFILE_DIR})…")
    # DETACHED_PROCESS + nowa grupa — Chrome przeżyje zamknięcie terminala.
    creationflags = 0x00000008 | 0x00000200  # DETACHED_PROCESS | CREATE_NEW_PROCESS_GROUP
    subprocess.Popen(args, close_fds=True, creationflags=creationflags)

    for _ in range(50):  # do ~10 s na rozruch
        if is_debug_chrome_running(port):
            return
        time.sleep(0.2)
    print(f"Chrome wystartował, ale port {port} nie odpowiada. Sprawdź czy jakiś inny Chrome nie blokuje.")
    sys.exit(1)


def build_driver(attach: bool = False) -> webdriver.Chrome:
    opts = Options()
    if attach:
        opts.debugger_address = f"127.0.0.1:{DEBUG_PORT}"
        return webdriver.Chrome(options=opts)

    # Ukrywamy wszystkie "developerskie" wskazowki w oknie Chrome.
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-infobars")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--no-first-run")
    opts.add_argument("--no-default-browser-check")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(options=opts)
    try:
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"},
        )
    except Exception:
        pass
    return driver


def ask_browser_mode() -> bool:
    """Zwraca True, jeśli pracujemy w trybie 'attach to Chrome'."""
    print("\nTryb przeglądarki:")
    print("  [1] Otwórz Chrome (skrypt sam go uruchomi) i podepnij się do niego — domyślnie")
    print("  [2] Uruchom nowe okno Chrome sterowane bezpośrednio przez Selenium")
    print("  [3] Podepnij się do już otwartego Chrome (jeżeli uruchomiłeś go wcześniej sam)")
    choice = input("Wybór [1]: ").strip() or "1"

    if choice == "2":
        return False

    if choice == "3":
        if not is_debug_chrome_running():
            print(f"\nNie widzę Chrome z DevTools na porcie {DEBUG_PORT}.")
            sys.exit(1)
        print("OK — podpinam się do działającego Chrome.")
        return True

    launch_debug_chrome()
    return True


def text_or_empty(driver, xpath: str) -> str:
    try:
        return driver.find_element(By.XPATH, xpath).text.strip()
    except NoSuchElementException:
        return ""


def get_field(driver, label: str) -> str:
    """Zwraca tekst z komórki <td> stojącej obok komórki z danym labelem."""
    xp = (
        f"//td[normalize-space()='{label}']/following-sibling::td[1]"
        f" | //th[normalize-space()='{label}']/following-sibling::td[1]"
    )
    return text_or_empty(driver, xp)


def semicolon_join(values: list[str]) -> str:
    return "; ".join(v.strip() for v in values if v and v.strip())


def scrape_dzial_io(driver) -> dict:
    """Scrapes Dział I-O (Oznaczenie nieruchomości) — najbardziej zmienny fragment EKW."""
    data = {c: "" for c in COLUMNS}

    data["Typ Księgi"] = get_field(driver, "Typ księgi wieczystej")
    data["Stan Księgi"] = get_field(driver, "Oznaczenie wydziału") or get_field(driver, "Stan")

    data["Województwo"] = get_field(driver, "Województwo")
    data["Powiat"] = get_field(driver, "Powiat")
    data["Gmina"] = get_field(driver, "Gmina")
    data["Miejscowość"] = get_field(driver, "Miejscowość")
    data["Dzielnica"] = get_field(driver, "Dzielnica")
    data["Położenie"] = get_field(driver, "Położenie")
    data["Ulica"] = get_field(driver, "Ulica")
    data["Sposób korzystania"] = get_field(driver, "Sposób korzystania")
    data["Obszar"] = get_field(driver, "Obszar")

    dzialki = [
        el.text.strip()
        for el in driver.find_elements(
            By.XPATH,
            "//td[contains(normalize-space(),'Numer działki')]/following-sibling::td[1]",
        )
    ]
    data["Nr działek po średniku"] = semicolon_join(dzialki)

    obreby = [
        el.text.strip()
        for el in driver.find_elements(
            By.XPATH,
            "//td[contains(normalize-space(),'Identyfikator działki')]/following-sibling::td[1]",
        )
    ]
    if not obreby:
        obreby = [
            el.text.strip()
            for el in driver.find_elements(
                By.XPATH,
                "//td[contains(normalize-space(),'Obręb ewidencyjny')]/following-sibling::td[1]",
            )
        ]
    data["Obręb po średniku"] = semicolon_join(obreby)

    data["Ulica(dla budynku)"] = get_field(driver, "Ulica (budynek)")
    data["przeznaczenie (dla budynku)"] = get_field(driver, "Przeznaczenie budynku")

    data["Ulica(dla lokalu)"] = get_field(driver, "Ulica (lokal)")
    data["Nr budynku( dla lokalu)"] = get_field(driver, "Numer budynku")
    data["Przeznaczenie (dla lokalu)"] = get_field(driver, "Przeznaczenie lokalu")
    data["Cały adres (dla lokalu)"] = get_field(driver, "Adres lokalu")

    udzialy = driver.find_elements(
        By.XPATH,
        "//*[contains(translate(., 'UDZIAŁ', 'udział'), 'udział')]",
    )
    data["Czy udziały?"] = "TAK" if udzialy else "NIE"

    return data


def random_delay(min_s: float = MIN_DELAY_SEC, max_s: float = MAX_DELAY_SEC) -> None:
    """Losowa przerwa — udaje człowieka i zmniejsza szansę na blokadę IP."""
    sec = random.uniform(min_s, max_s)
    print(f"    (odczekuję {sec:.1f}s, by nie zasypać serwera EKW)")
    time.sleep(sec)


def detect_block(driver) -> bool:
    """True, jeśli strona EKW wygląda na stronę blokady / limitu zapytań."""
    try:
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
    except NoSuchElementException:
        return False
    return any(phrase in body for phrase in BLOCK_PHRASES)


def batch_pause_if_needed(done: int) -> None:
    if done and done % BATCH_SIZE == 0:
        print(f"\n--- Wykonano {done} zapytań. Dłuższa przerwa {BATCH_PAUSE_SEC}s "
              f"(prewencja przed blokadą IP przez EKW). ---")
        time.sleep(BATCH_PAUSE_SEC)


def click_by_text(driver, *candidates: str, timeout: int = 15) -> bool:
    """Klika pierwszy znaleziony element zawierający jeden z podanych tekstów
    (link, button lub input submit). Zwraca True przy sukcesie."""
    end = time.time() + timeout
    while time.time() < end:
        for text in candidates:
            xp = (
                f"//a[contains(normalize-space(.), '{text}')]"
                f" | //button[contains(normalize-space(.), '{text}')]"
                f" | //input[@type='submit' and contains(@value, '{text}')]"
                f" | //input[@type='button' and contains(@value, '{text}')]"
            )
            els = driver.find_elements(By.XPATH, xp)
            for el in els:
                if el.is_displayed() and el.is_enabled():
                    try:
                        el.click()
                        return True
                    except Exception:
                        continue
        time.sleep(0.4)
    return False


def fill_and_wait(driver, kod: str, numer: str, cyfra: str) -> None:
    driver.get(EKW_URL)
    wait = WebDriverWait(driver, 20)
    wait.until(EC.presence_of_element_located((By.NAME, "kodWydzialuInput"))).clear()
    driver.find_element(By.NAME, "kodWydzialuInput").send_keys(kod)
    driver.find_element(By.NAME, "numerKsiegiWieczystej").clear()
    driver.find_element(By.NAME, "numerKsiegiWieczystej").send_keys(numer)
    driver.find_element(By.NAME, "cyfraKontrolna").clear()
    driver.find_element(By.NAME, "cyfraKontrolna").send_keys(cyfra)

    print("\n>>> Przepisz CAPTCHA w oknie przeglądarki i kliknij 'Szukaj księgi'.")
    input(">>> Gdy pojawi się strona z opcjami księgi — wciśnij ENTER tutaj: ")

    if detect_block(driver):
        raise RuntimeError(
            "EKW zwróciło stronę blokady / limitu zapytań. "
            "Przerwij pracę i poczekaj kilka/kilkanaście minut zanim spróbujesz ponownie."
        )

    print("    Klikam 'Przeglądanie aktualnej treści KW'…")
    if not click_by_text(
        driver,
        "Przeglądanie aktualnej treści KW",
        "aktualnej treści",
        "Przeglądanie aktualnej",
    ):
        raise RuntimeError(
            "Nie znaleziono przycisku 'Przeglądanie aktualnej treści KW'. "
            "Sprawdź czy jesteś na właściwej stronie po wyszukaniu księgi."
        )

    print("    Otwieram 'Dział I-O'…")
    WebDriverWait(driver, 15).until(
        lambda d: click_by_text(d, "Dział I-O", "Dzial I-O", "DZIAŁ I-O", "I-O", timeout=1)
    )

    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.XPATH, "//*[contains(., 'Oznaczenie')]"))
    )

    if detect_block(driver):
        raise RuntimeError("EKW zwróciło stronę blokady po otwarciu księgi.")


def row_already_done(ws, row: int, col_index: dict, kw_col: int) -> bool:
    """Wiersz traktujemy jako przetworzony, jeśli ma wypełnione dowolne
    pole wynikowe poza samym 'Nr KW'."""
    for name, idx in col_index.items():
        if name == "Nr KW" or idx == kw_col:
            continue
        val = ws.cell(row=row, column=idx).value
        if val not in (None, ""):
            return True
    return False


def main() -> None:
    xlsx_path = locate_or_create_workbook()

    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        kw_col = headers.index("Nr KW") + 1
    except ValueError:
        print("Brak kolumny 'Nr KW' w pierwszym wierszu.")
        sys.exit(1)

    col_index = {name: headers.index(name) + 1 for name in COLUMNS if name in headers}

    # --- wznawianie: policz, ile wierszy zostało do zrobienia -----------------
    already, todo = 0, 0
    for row in range(2, ws.max_row + 1):
        if not ws.cell(row=row, column=kw_col).value:
            continue
        if row_already_done(ws, row, col_index, kw_col):
            already += 1
        else:
            todo += 1
    print(f"\nPostęp w pliku: już wypełnione {already}, do zrobienia {todo}.")
    if todo == 0:
        print("Nic do zrobienia — wszystkie wiersze mają wypełnione dane.")
        return
    force_redo = False
    if already > 0:
        ans = input("Wznowić (T) czy przetworzyć wszystko od nowa i nadpisać (N)? [T/n]: ").strip().lower()
        if ans in ("n", "nie", "no"):
            force_redo = True
            print("OK — wszystkie wiersze zostaną przetworzone ponownie i nadpisane.")
        else:
            print("OK — pomijam wiersze już wypełnione.")
    # -------------------------------------------------------------------------

    attach = ask_browser_mode()
    driver = build_driver(attach=attach)
    done = 0
    try:
        for row in range(2, ws.max_row + 1):
            kw_value = ws.cell(row=row, column=kw_col).value
            if not kw_value:
                continue
            parts = split_kw(str(kw_value))
            if not parts:
                print(f"[wiersz {row}] Pomijam — nieprawidłowy format '{kw_value}' (oczekiwano XXXX/NNNNNNNN/C).")
                continue

            if not force_redo and row_already_done(ws, row, col_index, kw_col):
                print(f"[wiersz {row}] Już wypełniony — pomijam.")
                continue

            if done >= HARD_LIMIT_PER_RUN:
                print(f"\nOsiągnięto twardy limit {HARD_LIMIT_PER_RUN} zapytań w tym uruchomieniu.")
                print("Zakończam pracę, by nie narażać IP na blokadę. Uruchom skrypt ponownie później.")
                break

            batch_pause_if_needed(done)
            if done > 0:
                random_delay()

            kod, numer, cyfra = parts
            print(f"[wiersz {row}] KW: {kod}/{numer}/{cyfra}  (zapytanie {done + 1}/{HARD_LIMIT_PER_RUN})")

            try:
                fill_and_wait(driver, kod, numer, cyfra)
                data = scrape_dzial_io(driver)
            except RuntimeError as exc:
                print(f"[wiersz {row}] {exc}")
                break
            except Exception as exc:
                print(f"[wiersz {row}] Błąd pobierania danych: {exc}")
                done += 1
                continue

            data["Nr KW"] = f"{kod}/{numer}/{cyfra}"
            for name, idx in col_index.items():
                ws.cell(row=row, column=idx, value=data.get(name, ""))
            wb.save(xlsx_path)
            done += 1
            print(f"[wiersz {row}] Zapisano.")
    finally:
        if attach:
            # Nie zamykamy Chrome użytkownika — odłączamy się tylko od sesji.
            try:
                driver.service.stop()
            except Exception:
                pass
        else:
            driver.quit()

    print(f"\nGotowe. Wykonano {done} zapytań. Plik: {xlsx_path}")


if __name__ == "__main__":
    main()
